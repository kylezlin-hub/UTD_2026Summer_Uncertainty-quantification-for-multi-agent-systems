import re
from typing import Dict, List, Optional

from openpyxl import load_workbook


def _safe_rate(numerator: int, denominator: int) -> Optional[float]:
    """Return numerator / denominator, or None when the rate is undefined."""
    if denominator == 0:
        return None
    return numerator / denominator


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_answer(value: object) -> Optional[str]:
    text = _clean_text(value)
    if not text:
        return None
    match = re.search(r"\b([A-J])\b", text.upper())
    if match:
        return match.group(1)
    if len(text) == 1 and text.upper().isalpha():
        return text.upper()
    return re.sub(r"\s+", " ", text)


def _is_yes(value: object) -> Optional[bool]:
    text = _clean_text(value).lower()
    if text in {"yes", "y", "true", "1", "correct"}:
        return True
    if text in {"no", "n", "false", "0", "incorrect"}:
        return False
    return None


def accuracy(correct_final_answers: int, total_questions: int) -> Optional[float]:
    """Calculate accuracy as correct final answers divided by total questions."""
    return _safe_rate(correct_final_answers, total_questions)






def consensus_reliability(
    correct_consensus_answers: int,
    total_consensus_questions: int,
) -> Optional[float]:
    """Calculate Prob(correct | consensus).

    This answers: when agents agree, how often are they actually right?
    """
    return _safe_rate(correct_consensus_answers, total_consensus_questions)


def premature_elimination_rate(
    correct_hypotheses_discarded_before_final_rounds: int,
    all_correct_hypotheses_proposed: int,
) -> Optional[float]:
    """Calculate correct hypotheses discarded early divided by all proposed.

    Example: if correct answers were proposed early 70 times and discarded
    early 28 times, PER = 28 / 70 = 0.40.
    """
    return _safe_rate(
        correct_hypotheses_discarded_before_final_rounds,
        all_correct_hypotheses_proposed,
    )


Premature_Elimination_Rate = premature_elimination_rate


def load_debate_rows(
    excel_path: str,
    sheet_name: str = "Debate_Traces",
) -> List[Dict[str, object]]:
    """Load debate rows from an Excel workbook into dictionaries."""
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    headers = list(next(rows))
    return [dict(zip(headers, row)) for row in rows]


def get_round_answers(
    row: Dict[str, object],
    round_no: int,
    agent_count: int = 3,
) -> List[Optional[str]]:
    """Return normalized agent answers for one debate round."""
    return [
        _normalize_answer(row.get(f"R{round_no} Agent{agent_no} Answer"))
        for agent_no in range(1, agent_count + 1)
    ]


def final_answer_is_correct(row: Dict[str, object]) -> bool:
    """Return whether the final answer is correct for one debate row."""
    labeled_correctness = _is_yes(row.get("Correct?"))
    if labeled_correctness is not None:
        return labeled_correctness
    return _normalize_answer(row.get("Final Answer")) == _normalize_answer(
        row.get("Correct Answer")
    )


def agents_agree(answers: List[Optional[str]]) -> bool:
    """Return True when all non-empty answers are the same."""
    non_empty_answers = [answer for answer in answers if answer is not None]
    return bool(non_empty_answers) and len(set(non_empty_answers)) == 1


def row_has_consensus(
    row: Dict[str, object],
    consensus_mode: str = "final_round",
    final_round: int = 5,
    agent_count: int = 3,
) -> bool:
    """Return whether a debate row reached consensus.

    consensus_mode options:
    - "final_round": agents agree in the final round.
    - "source": Final Answer Source is agent_consensus.
    - "rounds_to_consensus": Rounds to Consensus is non-empty.
    """
    if consensus_mode == "final_round":
        return agents_agree(get_round_answers(row, final_round, agent_count))
    if consensus_mode == "source":
        return _clean_text(row.get("Final Answer Source")) == "agent_consensus"
    if consensus_mode == "rounds_to_consensus":
        rounds_to_consensus = _clean_text(row.get("Rounds to Consensus"))
        return bool(rounds_to_consensus) and rounds_to_consensus.lower() != "nan"
    raise ValueError(f"Unknown consensus_mode: {consensus_mode}")


def correct_answer_proposed_before_final(
    row: Dict[str, object],
    final_round: int = 5,
    agent_count: int = 3,
) -> bool:
    """Return True when any agent proposed the correct answer before final round."""
    correct_answer = _normalize_answer(row.get("Correct Answer"))
    early_answers = [
        answer
        for round_no in range(1, final_round)
        for answer in get_round_answers(row, round_no, agent_count)
    ]
    return correct_answer is not None and correct_answer in early_answers


def correct_answer_discarded_by_final(
    row: Dict[str, object],
    final_round: int = 5,
    agent_count: int = 3,
) -> bool:
    """Return True when an early correct answer is absent from the final round."""
    correct_answer = _normalize_answer(row.get("Correct Answer"))
    if not correct_answer_proposed_before_final(row, final_round, agent_count):
        return False
    return correct_answer not in get_round_answers(row, final_round, agent_count)


def calculate_debate_metrics(
    excel_path: str,
    sheet_name: str = "Debate_Traces",
    final_round: int = 5,
    agent_count: int = 3,
    consensus_mode: str = "final_round",
) -> Dict[str, Optional[float]]:
    """Calculate debate metrics from a workbook.

    Premature elimination is debate-level: each question counts once if the
    correct answer appeared before the final round, and counts as discarded if
    no final-round agent kept that correct answer.
    """
    rows = load_debate_rows(excel_path, sheet_name)

    total_questions = len(rows)
    correct_final_answers = sum(final_answer_is_correct(row) for row in rows)

    consensus_rows = [
        row
        for row in rows
        if row_has_consensus(row, consensus_mode, final_round, agent_count)
    ]
    correct_consensus_answers = sum(final_answer_is_correct(row) for row in consensus_rows)

    correct_hypotheses_proposed = sum(
        correct_answer_proposed_before_final(row, final_round, agent_count)
        for row in rows
    )
    correct_hypotheses_discarded = sum(
        correct_answer_discarded_by_final(row, final_round, agent_count)
        for row in rows
    )

    return {
        "total_questions": total_questions,
        "correct_final_answers": correct_final_answers,
        "accuracy": accuracy(correct_final_answers, total_questions),
        "consensus_questions": len(consensus_rows),
        "correct_consensus_answers": correct_consensus_answers,
        "consensus_reliability": consensus_reliability(
            correct_consensus_answers,
            len(consensus_rows),
        ),
        "correct_hypotheses_proposed_before_final": correct_hypotheses_proposed,
        "correct_hypotheses_discarded_before_final": correct_hypotheses_discarded,
        "premature_elimination_rate": premature_elimination_rate(
            correct_hypotheses_discarded,
            correct_hypotheses_proposed,
        ),
    }


if __name__ == "__main__":
    metrics = calculate_debate_metrics(
        "data/qwen_mmlu_pro_200/qwen_mmlu_pro_debate_traces.xlsx"
    )
    for name, value in metrics.items():
        print(f"{name}: {value}")
